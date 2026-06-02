# -*- coding: utf-8 -*-
"""Сервис экспорта данных в CSV и Excel."""

import csv
import io
import os
import re
from typing import List, Optional, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from fastapi.responses import FileResponse, Response

from app.config import settings
from app.models.domain import FirmData


EXPORT_HEADERS = [
    'Название', 'Оценка', 'Город', 'Адрес', 'Расписание', 'Телефоны', 'Почта',
    'Сайт',
    'Whatsapp номер 1', 'Whatsapp номер 2', 'Whatsapp ссылка',
    'Telegram ссылка', 'Telegram ник',
    'Instagram', 'Youtube', 'VK', 'OK', 'Другие соцсети',
    'Информация', 'URL',
]


def normalize_export_columns(columns: Optional[Sequence[str]]) -> List[str]:
    """Валидация и упорядочивание выбранных колонок."""
    if not columns:
        return list(EXPORT_HEADERS)

    allowed = set(EXPORT_HEADERS)
    selected = [c.strip() for c in columns if c and c.strip() in allowed]
    if not selected:
        raise ValueError('Не выбрано ни одной допустимой колонки')

    return [h for h in EXPORT_HEADERS if h in selected]


class CsvExporter:
    """Экспорт данных в CSV (UTF-8 с BOM для Excel)."""

    def __init__(self):
        self.export_dir = Path(settings.EXCEL_DIR)
        self.export_dir.mkdir(exist_ok=True)

    def export_to_file(
        self,
        task_id: str,
        firms: List[FirmData],
        columns: Optional[Sequence[str]] = None,
    ) -> str:
        """Экспорт фирм в CSV файл."""
        filename = f"2gis_export_{task_id}.csv"
        filepath = self.export_dir / filename
        fieldnames = normalize_export_columns(columns)
        content = self.build_csv_bytes(firms, fieldnames)

        with open(filepath, 'wb') as f:
            f.write(content)

        return str(filepath)

    def build_csv_bytes(
        self,
        firms: List[FirmData],
        columns: Optional[Sequence[str]] = None,
    ) -> bytes:
        """Сборка CSV в памяти с выбранными колонками."""
        fieldnames = normalize_export_columns(columns)
        rows = self._build_unique_whatsapp_rows(firms)

        buffer = io.StringIO()
        writer = csv.DictWriter(
            buffer, fieldnames=fieldnames, extrasaction='ignore'
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

        return buffer.getvalue().encode('utf-8-sig')

    def _build_unique_whatsapp_rows(self, firms: List[FirmData]) -> List[dict]:
        """
        Подготовка строк для экспорта:
        - WhatsApp 1/2 уже ограничены на уровне карточки;
        - повторы номеров между строками очищаются.
        """
        rows: List[dict] = []
        used_whatsapp_numbers = set()

        for firm in firms:
            row = firm.to_dict()

            for key in ('Whatsapp номер 1', 'Whatsapp номер 2'):
                raw = str(row.get(key, '') or '').strip()
                if not raw:
                    continue

                # Нормализуем в цифры, чтобы +7..., 7... и 77... считались одинаковыми.
                normalized = re.sub(r'\D+', '', raw)
                number_key = normalized or raw

                if number_key in used_whatsapp_numbers:
                    row[key] = ''
                    continue

                used_whatsapp_numbers.add(number_key)
                row[key] = number_key

            rows.append(row)

        return rows

    def export_urls(self, task_id: str, urls: List[str]) -> str:
        """Экспорт списка URL организаций в CSV."""
        filename = f"2gis_urls_{task_id}.csv"
        filepath = self.export_dir / filename

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['URL'])
            for url in urls:
                writer.writerow([url])

        return str(filepath)

    def get_file_response(
        self,
        filepath: str,
        filename: Optional[str] = None,
    ) -> FileResponse:
        """FileResponse для скачивания CSV."""
        return FileResponse(
            path=filepath,
            filename=filename or os.path.basename(filepath),
            media_type='text/csv; charset=utf-8',
        )

    def get_bytes_response(self, content: bytes, filename: str) -> Response:
        """Скачивание CSV, сгенерированного в памяти."""
        return Response(
            content=content,
            media_type='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
            },
        )


class ExcelExporter:
    """Класс для экспорта данных в Excel."""
    
    HEADERS = EXPORT_HEADERS
    
    COLUMN_WIDTHS = {
        'Название': 30, 'Оценка': 10, 'Город': 18, 'Адрес': 35, 'Расписание': 25,
        'Телефоны': 25, 'Почта': 30, 'Сайт': 40,
        'Whatsapp номер 1': 18, 'Whatsapp номер 2': 18, 'Whatsapp ссылка': 40,
        'Telegram ссылка': 35, 'Telegram ник': 18,
        'Instagram': 40, 'Youtube': 40, 'VK': 35, 'OK': 35,
        'Другие соцсети': 45, 'Информация': 50,
        'URL': 45,
    }
    
    def __init__(self):
        self.export_dir = Path(settings.EXCEL_DIR)
        self.export_dir.mkdir(exist_ok=True)
    
    def export_to_file(self, task_id: str, firms: List[FirmData]) -> str:
        """Экспорт фирм в Excel файл."""
        filename = f"2gis_export_{task_id}.xlsx"
        filepath = self.export_dir / filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"
        
        # Заголовки
        ws.append(self.HEADERS)
        
        # Установка ширины столбцов
        for idx, header in enumerate(self.HEADERS, 1):
            col_letter = ws.cell(row=1, column=idx).column_letter
            width = self.COLUMN_WIDTHS.get(header, 20)
            ws.column_dimensions[col_letter].width = width
        
        # Данные (с той же уникализацией WhatsApp, что и в CSV)
        rows = CsvExporter()._build_unique_whatsapp_rows(firms)
        for row in rows:
            values = [row.get(h, '') for h in self.HEADERS]
            ws.append(values)
            
            row_idx = ws.max_row
            for col_idx in range(1, len(self.HEADERS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(
                    wrap_text=False,
                    vertical='center',
                    shrink_to_fit=False
                )
            ws.row_dimensions[row_idx].height = 15
        
        # Создание умной таблицы
        if ws.max_row > 1:
            last_col = ws.cell(row=1, column=len(self.HEADERS)).column_letter
            table_range = f"A1:{last_col}{ws.max_row}"
            
            table = Table(displayName="DataTable", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(table)
        
        wb.save(filepath)
        return str(filepath)
    
    def get_file_response(self, filepath: str) -> FileResponse:
        """Получение FileResponse для скачивания."""
        return FileResponse(
            path=filepath,
            filename=os.path.basename(filepath),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def delete_file(self, filepath: str) -> None:
        """Удаление файла."""
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception:
            pass


# Глобальные экземпляры экспортёров
csv_exporter = CsvExporter()
exporter = csv_exporter  # основной формат выгрузки — CSV
excel_exporter = ExcelExporter()

